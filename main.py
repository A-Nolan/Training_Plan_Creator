from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl import load_workbook

MANAGERS = [
    'Aaron Nolan',
    'Laurence Murphy',
    'Sebastian Staszak',
    'Nicole Kelly',
    'Mellissa Kelly',
    'Dumitru Popa',
    'aoife Byrne',
    'Joe Dudley',
    'Katarzyna Hajrych',
    'Stephanie Rooney',
    'Matthew Taylor',
    'Edyta Zawada'
]

CREW_TRAINERS = [
    'Caoimhe Kelly',
    'Emma Power',
    'Barry Doyle',
    'Nicole O\'Brien',
    'Lauren Moore',
    'Martell Cullen',
    'Robyn Moore Keogh',
    'William Nolan',
    'Raphael Vieira'
]

NOT_RETURNED = [
    'Michelle Buchanan Culava Toma',
    'Dylan Clear',
    'James Haughton Kellett',
    'Filip Michalski',
    'Catherine Murphy',
    'Sofia Rooney',
    'Jadwiga Sosnowska',
    'Tomasz Stefaniak',
    'William Swayne',
    'Monika Sykta'
]

# METHODS TO CREATE A .XLSX FROM MYSCHEDULE
def schedule_to_excel():
    #input users id and password
    user = input('User ID: ')
    password = input('Password: ')
    week_num = input('Week Number: ')

    driver = open_schedule(user, password, week_num)
    schedule_td = driver.find_elements_by_css_selector('#gridbox > table > tbody > tr:nth-child(2) > td > div > div > table > tbody > tr > td')

    print('schedule.xlsx building ....')

    index = 0

    schedule_wb = Workbook()
    schedule_ws = schedule_wb.active

    for col in range(1, (int(len(schedule_td) / 11)) + 1):
        for row in range(1, 12):
            info = schedule_td[index].text
            if row == 1:
                name = rearrange_name(info)
                info = name
            curr_cell = schedule_ws.cell(col, row)
            curr_cell.value = info
            index += 1

    add_suggested_socs(schedule_ws)

    schedule_wb.save('schedule.xlsx')
    print('schedule.xlsx created successfully')

    wb2 = Workbook()
    wb2.create_sheet('Crew Members')
    wb2.create_sheet('Crew Trainers')
    wb2.create_sheet('Managers')
    wb2.create_sheet('Not Returned')
    ws = wb2.active
    wb2.remove(ws)
    wb2.save('training_planner.xlsx')

def open_schedule(user, password, week_num):
    """ Open Selenium Link to the schedule

    :Args:
    - `str`: `user` - Username for MySchedule
    - `str`: `password` - Password for MySchedule
    - `str`: `week_num` - The week number of the schedule
    """

    options = Options()
    options.add_argument('headless')

    driver = webdriver.Chrome(options=options, executable_path='C:\Program Files (x86)\chromedriver')
    driver.get('https://psschedule.reflexisinc.co.uk/wfmmcdirlprd/ModuleSelection.jsp')

    # Log in
    userid_input_box = driver.find_element_by_name("txtUserID")
    password_input_box = driver.find_element_by_name('txtPassword')
    userid_input_box.send_keys(user)
    password_input_box.send_keys(password)
    login_button = driver.find_element_by_class_name('button-t')
    login_button.click()

    driver.get(f'https://psschedule.reflexisinc.co.uk/wfmmcdirlprd/rws/schedule/schedule_weekly.jsp?sm=0&mm=SCHD&cboYear=2020&cboQuarter=3&cboMonth=7&cboWeek={week_num}&showYearDropDown=Y')

    return driver

def rearrange_name(name):
    """ Remove the comma and reverse the name from MySchedule

    :Args:
    - `str`: `name` - The wrongly formatted name `Nolan, Aaron`

    :Returns:
    - `str` - The reformatted name `Aaron Nolan`
    """

    name_l = name.split()
    name_l.remove(',')
    name_l.insert(0, name_l.pop())
    return ' '.join(name_l)

def add_suggested_socs(schedule_ws):
    """ Add the suggested SOCs to the schedule

    :Args:
    - `openpyxl.Worksheet`: `schedule_ws` - The Schedule Worksheet
    """

    sugg_soc_wb = load_workbook('Suggested_SOCs.xlsx')
    sugg_soc_ws = sugg_soc_wb.active

    soc_dict = {}

    for row in sugg_soc_ws.iter_rows(min_col=2, max_col=5, min_row=2, values_only=True):
        soc_dict[row[0]] = [row[1], row[2], row[3]]

    for index, row in enumerate(schedule_ws.iter_rows(values_only=True), start=1):
        if row[0] in soc_dict.keys():
            schedule_ws.cell(row=index, column=12).value = soc_dict[row[0]][0]
            schedule_ws.cell(row=index, column=13).value = soc_dict[row[0]][1]
            schedule_ws.cell(row=index, column=14).value = soc_dict[row[0]][2]

# METHODS TO CREATE TRAINING_PLANNER.XLSX
def schedule_to_planner():
    print('training_planner.xlsx building ....')

    schedule_wb = load_workbook('schedule.xlsx')
    planner_wb = load_workbook('training_planner.xlsx')

    schedule_ws = schedule_wb.active
    cm_planner_ws = planner_wb['Crew Members']
    ct_planner_ws = planner_wb['Crew Trainers']
    man_planner_ws = planner_wb['Managers']
    nr_planner_ws = planner_wb['Not Returned']

    cm_planner_ws.append(['Name', '1', '2', '3', '4', '5', '6', '7', 'SOC 1', 'SOC 2', 'SOC 3', 'Completed SOCs'])

    for row in schedule_ws.iter_rows(values_only=True):
        if row[0] in NOT_RETURNED:
            nr_planner_ws.append([row[0], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13]])
        elif row[0] in CREW_TRAINERS:
            ct_planner_ws.append([row[0], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13]])
        elif row[0] in MANAGERS:
            man_planner_ws.append([row[0], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13]])
        else:
            cm_planner_ws.append([row[0], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13]])

    complete_worksheet(cm_planner_ws)
    complete_worksheet(ct_planner_ws)
    complete_worksheet(man_planner_ws)
    add_soc_count(cm_planner_ws)

    planner_wb.save('training_planner.xlsx')
    print('training_planner.xlsx created successfully')

def complete_worksheet(ws):
    """ Formats `schedule.xlsx` into start times and blanks

    :Args:
    - `openpyxl.Worksheet`: `ws` - Schedule Worksheet
    """

    for row in ws.iter_rows(min_col=2, max_col=8):
        for cell in row:
            if cell.value == None or not cell.value[0].isnumeric():
                cell.value = None
            else:
                cell.value = cell.value[:2] + cell.value[3:5]
                cell.value = int(cell.value)

def add_soc_count(ws):
    # On PC C:\\Users\\Aaron\\OneDrive\\Work\\Learning & Development\\Training Plan.xlsx
    # on Laptop C:\\Users\\Xtrem\\OneDrive\\Work\\Learning & Development\\Training Plan.xlsx
    training_plan_wb = load_workbook('C:\\Users\\Xtrem\\OneDrive\\Work\\Learning & Development\\Training Plan.xlsx')
    training_plan_ws = training_plan_wb['Crew SOC Count']

    count_dict = {}

    for row in training_plan_ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] not in CREW_TRAINERS and row[0] not in MANAGERS and row[0] not in NOT_RETURNED and row[0] != None:
            count_dict[row[0]] = row[1]

    for index, row in enumerate(ws.iter_rows(), start=1):
        if index == 1:
            pass
        else:
            ws.cell(index, 12).value = count_dict[row[0].value]

# METHODS TO CREATE THE TRAINING PLAN
def create_training_plan():
    training_planner_wb = load_workbook('training_planner.xlsx')
    training_planner_wb.create_sheet('Training Plan')

    cm_ws = training_planner_wb['Crew Members']
    ct_ws = training_planner_wb['Crew Trainers']
    tp_ws = training_planner_wb['Training Plan']

    # How many have been planned for the week
    planned_socs = 0

    # what number are we checking for completions
    completed_socs = 0

    while planned_socs < 7:
        for index, row in enumerate(cm_ws.iter_rows(values_only=True), start=1):

            max_row = cm_ws.max_row

            if planned_socs > 6:
                break
            if row[len(row) - 1] == completed_socs:
                day = choose_day_from_row(row, cm_ws)
                if day != 0:
                    add_to_plan(row, tp_ws, day)
                    planned_socs += 1
                    cm_ws.delete_rows(index, 1)
                    col_to_delete = find_col_to_delete(cm_ws[1], day)
                    cm_ws.delete_cols(col_to_delete, 1)
                    break
            if index > max_row - 1:
                completed_socs += 1
        
    add_cts_to_plan(ct_ws, tp_ws)

    training_planner_wb.save('training_planner.xlsx')

def choose_day_from_row(row, ws):
    """ Return what day that person can have an SOC

    :Args:
    - `list`: `row` - List of info pulled from a row in excel

    :Returns:
    - `int` - (1-7)(Mon - Fri) or 0, No available dates
    """

    for index in range(2, len(row) - 3):
        if row[index - 1] == '' or row[index - 1] == None:
            pass
        else:
            return int(ws.cell(1, index).value)

    return 0

def add_to_plan(row, ws, day):
    days = [
        'Monday',
        'Tuesday',
        'Wednesday',
        'Thursday',
        'Friday',
        'Saturday',
        'Sunday'
    ]

    # Add day to first column
    ws.cell(day, 1, days[day - 1])

    # Add name to 2nd column
    ws.cell(day, 2, row[0])

    # Add SOC to 4th Column
    ws.cell(day, 4, row[len(row) - 4])

    # Add start time to 5th Column
    for index, cell in enumerate(row):
        if cell != None and index > 0:
            time = row[index]
            break
    ws.cell(day, 5, time)

    print(f'{days[day - 1]}, {row[0]}, {row[len(row) - 4]}, {time}')

def find_col_to_delete(row, day):
    for index, cell in enumerate(row):
        if cell.value == day:
            return index + 1

def add_cts_to_plan(ct_ws, tp_ws):

    closest_time = 2400

    for index, ct_col in enumerate(ct_ws.iter_cols(min_col=1, max_col=8, values_only=True), 1):
        
        if index > 1:
            cm_time = int(tp_ws[index - 1][4].value)
            for index2, time in enumerate(ct_col, 1):
                if time != None:
                    ct_time = int(ct_ws.cell(index2, index).value)
                    time_diff = abs(cm_time - ct_time)
                    if time_diff < closest_time:
                        tp_ws[index - 1][2].value = ct_ws[index2][0].value
                        closest_time = time_diff

            closest_time = 2400

# schedule_to_excel()
# add_suggested_socs()
# schedule_to_planner()
# create_training_plan()

test_wb = load_workbook('training_planner.xlsx')
wb_to_copy = test_wb['Crew Members']
test_wb.copy_worksheet(wb_to_copy)
cm_foundation_ws = test_wb['Crew Members Copy']
test_wb.move_sheet(cm_foundation_ws, -4)
cm_foundation_ws.title = 'Crew Member Foundation'
print(test_wb.sheetnames)